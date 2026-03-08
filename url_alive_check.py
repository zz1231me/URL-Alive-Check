#!/usr/bin/env python3
"""
URL_Alive_Check - URL 접속 상태 확인 도구
- Playwright를 사용하여 Direct / Proxy 시나리오 테스트
- 프록시 사용/미사용 각각 테스트
- URL 변형(http/https/www) 자동 테스트
- 결과를 XLSX(URL당 1행)로 저장
"""

import os
import time
import csv
import json
import argparse
import sys
import threading
from datetime import datetime
from typing import Dict, List, Optional
from urllib.parse import urlparse
from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeout
from colorama import Fore, Style, init
try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
    _OPENPYXL_AVAILABLE = True
except ImportError:
    _OPENPYXL_AVAILABLE = False

# Colorama 초기화
init(autoreset=True)


class URLAliveCheck:
    """URL 접속 상태 확인 클래스"""

    def __init__(self, config_file: Optional[str] = None):
        """
        초기화

        Args:
            config_file: 설정 파일 경로 (JSON)
        """
        self.config = self._load_config(config_file)
        self.results = []

    @staticmethod
    def _deep_merge(base: dict, override: dict) -> dict:
        """딕셔너리를 재귀적으로 깊은 병합"""
        result = base.copy()
        for key, value in override.items():
            if key in result and isinstance(result[key], dict) and isinstance(value, dict):
                result[key] = URLAliveCheck._deep_merge(result[key], value)
            else:
                result[key] = value
        return result

    def _load_config(self, config_file: Optional[str]) -> Dict:
        """설정 파일 로드"""
        default_config = {
            'proxy': {
                'enabled': False,
                'server': 'http://proxy-server.com:8080',
                'use_auth': False,
                'username': '',
                'password': ''
            },
            'timeout': 30,
            'headless': True,
            'user_agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
            'retry_count': 0,
            'retry_delay': 5
        }

        if config_file:
            try:
                with open(config_file, 'r', encoding='utf-8') as f:
                    user_config = json.load(f)
                    default_config = self._deep_merge(default_config, user_config)
            except FileNotFoundError:
                print(f"  {Fore.YELLOW}▲  설정 파일을 찾을 수 없습니다. 기본 설정을 사용합니다.")
            except json.JSONDecodeError:
                print(f"  {Fore.RED}✕  설정 파일 형식이 올바르지 않습니다. 기본 설정을 사용합니다.")

        return default_config

    @staticmethod
    def _normalize_url(url: str) -> Optional[str]:
        """URL에 스킴이 없으면 https:// 자동 추가. 빈 문자열이면 None 반환"""
        url = url.strip()
        if not url:
            return None
        if not url.startswith(('http://', 'https://')):
            url = 'https://' + url
        return url

    @staticmethod
    def _generate_url_variations(url: str) -> List[str]:
        """URL의 http/https, www/non-www 변형 생성"""
        parsed = urlparse(url)
        hostname = parsed.hostname or ''

        # www 제거한 기본 호스트
        if hostname.startswith('www.'):
            base_host = hostname[4:]
        else:
            base_host = hostname

        # 경로 부분 재구성
        path_portion = parsed.path or ''
        if parsed.query:
            path_portion += '?' + parsed.query
        if parsed.fragment:
            path_portion += '#' + parsed.fragment

        # 포트 처리
        port_suffix = ''
        if parsed.port and parsed.port not in (80, 443):
            port_suffix = f':{parsed.port}'

        variations = [
            f'https://{base_host}{port_suffix}{path_portion}',
            f'http://{base_host}{port_suffix}{path_portion}',
            f'https://www.{base_host}{port_suffix}{path_portion}',
            f'http://www.{base_host}{port_suffix}{path_portion}',
        ]

        # 중복 제거 (순서 유지)
        seen = set()
        unique = []
        for v in variations:
            if v not in seen:
                seen.add(v)
                unique.append(v)
        return unique

    _print_lock = threading.Lock()

    def _log(self, message: str, end='\n'):
        """실시간 콘솔 출력 (스레드 안전)"""
        with URLAliveCheck._print_lock:
            print(message, end=end, flush=True)

    def _get_playwright_proxy_config(self) -> Optional[Dict]:
        """Playwright용 프록시 설정 반환"""
        if not self.config['proxy']['enabled']:
            return None

        proxy_config = {
            'server': self.config['proxy']['server']
        }

        use_auth = self.config['proxy'].get('use_auth', False)

        if use_auth:
            username = self.config['proxy'].get('username', '')
            password = self.config['proxy'].get('password', '')

            if not username or not password:
                self._log(f"  {Fore.RED}✕  프록시 인증 정보가 불완전합니다 (username 또는 password 누락).")
                self._log(f"  {Fore.YELLOW}▲  프록시를 비활성화하고 Direct 모드로만 테스트합니다.")
                return None

            proxy_config['username'] = username
            proxy_config['password'] = password

        return proxy_config

    @staticmethod
    def check_playwright_available() -> bool:
        """Playwright 및 Chromium 브라우저 로딩 가능 여부 사전 검증"""
        W = 68
        print(f"{Fore.CYAN}┌{'─'*(W-2)}┐")
        print(f"{Fore.CYAN}│{Style.RESET_ALL}  Playwright 환경 검증")
        print(f"{Fore.CYAN}├{'─'*(W-2)}┤", flush=True)

        # 1) playwright 패키지 import 확인
        try:
            from playwright.sync_api import sync_playwright  # noqa: F401
            print(f"{Fore.CYAN}│{Style.RESET_ALL}  {Fore.GREEN}●{Style.RESET_ALL}  playwright 패키지 로드 성공")
        except ImportError:
            print(f"{Fore.CYAN}│{Style.RESET_ALL}  {Fore.RED}✕{Style.RESET_ALL}  playwright 패키지를 찾을 수 없습니다.")
            print(f"{Fore.CYAN}│{Style.RESET_ALL}     {Fore.YELLOW}→ pip install playwright 실행 후 재시도하세요.")
            print(f"{Fore.CYAN}└{'─'*(W-2)}┘")
            return False

        # 2) Chromium 브라우저 실행 확인
        try:
            with sync_playwright() as p:
                browser = p.chromium.launch(headless=True)
                version = browser.version
                browser.close()
            print(f"{Fore.CYAN}│{Style.RESET_ALL}  {Fore.GREEN}●{Style.RESET_ALL}  Chromium 실행 성공  {Style.DIM}(버전: {version}){Style.RESET_ALL}")
        except Exception as e:
            err_msg = str(e)
            print(f"{Fore.CYAN}│{Style.RESET_ALL}  {Fore.RED}✕{Style.RESET_ALL}  Chromium 실행 실패:")
            # 오류 메시지를 줄당 W-6자씩 나눠 출력
            chunk = W - 6
            for i in range(0, len(err_msg), chunk):
                print(f"{Fore.CYAN}│{Style.RESET_ALL}     {Fore.RED}{err_msg[i:i+chunk]}{Style.RESET_ALL}")
            print(f"{Fore.CYAN}│{Style.RESET_ALL}     {Fore.YELLOW}→ playwright install chromium 실행 후 재시도하세요.")
            print(f"{Fore.CYAN}└{'─'*(W-2)}┘")
            return False

        print(f"{Fore.CYAN}│{Style.RESET_ALL}  {Fore.GREEN}●{Style.RESET_ALL}  검증 완료 — 모니터링을 시작합니다.")
        print(f"{Fore.CYAN}└{'─'*(W-2)}┘\n")
        return True

    def _test_page_with_browser(self, browser, url: str) -> Dict:
        """이미 열린 browser 인스턴스로 페이지 테스트 (내부용)"""
        result = {
            'method': 'playwright',
            'status': None,
            'http_code': None,
            'response_time': None,
            'error_message': None,
            'final_url': None,
            'console_errors': [],
            'network_errors': [],
            'redirect_chain': []
        }
        timeout = self.config['timeout'] * 1000

        try:
            context = browser.new_context(
                user_agent=self.config['user_agent'],
                viewport={'width': 1920, 'height': 1080},
                locale='ko-KR',
                timezone_id='Asia/Seoul',
                extra_http_headers={
                    'Accept-Language': 'ko-KR,ko;q=0.9,en-US;q=0.8,en;q=0.7',
                    'Accept-Encoding': 'gzip, deflate, br',
                    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8'
                }
            )
            try:
                page = context.new_page()

                # navigator.webdriver 속성 제거 (봇 감지 우회)
                page.add_init_script("""
                    Object.defineProperty(navigator, 'webdriver', {
                        get: () => undefined
                    });
                """)

                def handle_console(msg):
                    if msg.type == 'error':
                        result['console_errors'].append(msg.text)

                page.on('console', handle_console)

                def handle_request_failed(request):
                    # 전체 URL 대신 scheme+host만 저장해 내부 경로/파라미터 노출 방지
                    try:
                        parsed = urlparse(request.url)
                        safe_url = f"{parsed.scheme}://{parsed.netloc}"
                    except Exception:
                        safe_url = '(unknown)'
                    result['network_errors'].append(safe_url)

                page.on('requestfailed', handle_request_failed)

                redirects = []

                def handle_response(resp):
                    if resp.status in (301, 302, 303, 307, 308):
                        redirects.append({'code': resp.status, 'url': resp.url})

                page.on('response', handle_response)

                start_time = time.time()
                response = page.goto(url, timeout=timeout, wait_until='domcontentloaded')

                try:
                    page.wait_for_load_state('networkidle', timeout=timeout)
                except PlaywrightTimeout:
                    pass

                response_time = time.time() - start_time

                result['http_code'] = response.status if response else None
                result['response_time'] = round(response_time, 3)
                result['final_url'] = page.url
                result['redirect_chain'] = redirects

                if response and 200 <= response.status < 300:
                    was_redirected = len(redirects) > 0 or (page.url != url)
                    if was_redirected:
                        result['status'] = 'REDIRECT'
                        result['error_message'] = f"Redirected: {url} -> {page.url}"
                    elif result['console_errors']:
                        result['status'] = 'WARNING'
                        result['error_message'] = f"JavaScript errors: {len(result['console_errors'])} errors"
                    elif result['network_errors']:
                        result['status'] = 'WARNING'
                        result['error_message'] = f"Network errors: {len(result['network_errors'])} failed"
                    else:
                        result['status'] = 'SUCCESS'
                elif response and 400 <= response.status < 500:
                    result['status'] = 'ERROR'
                    result['error_message'] = f"Client error: {response.status}"
                elif response and 500 <= response.status < 600:
                    result['status'] = 'ERROR'
                    result['error_message'] = f"Server error: {response.status}"
                else:
                    result['status'] = 'ERROR'
                    result['error_message'] = 'Unknown error'
            finally:
                context.close()

        except PlaywrightTimeout:
            result['status'] = 'TIMEOUT'
            result['error_message'] = 'Navigation timeout'
            result['response_time'] = self.config['timeout']
        except Exception as e:
            result['status'] = 'ERROR'
            result['error_message'] = f'Playwright error: {str(e)}'

        return result

    @staticmethod
    def _fmt_time(t) -> str:
        """응답시간을 고정폭 6자 문자열로 포맷"""
        return f"{t:>6.2f}s" if t is not None else "      -"

    def _test_single_variation(self, browser, proxy_browser, url: str) -> Dict:
        """이미 열린 browser(direct)와 proxy_browser로 하나의 URL 변형을 테스트 (출력 없음)"""
        results = {
            'url': url,
            'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'direct': {'playwright': None},
            'proxy': {'playwright': None}
        }

        retry_count = self.config.get('retry_count', 0)
        retry_delay = self.config.get('retry_delay', 5)

        def _test_with_retry(b) -> Dict:
            last = {
                'method': None, 'status': 'ERROR', 'http_code': None,
                'response_time': None, 'error_message': 'No result returned',
                'final_url': None, 'console_errors': [], 'network_errors': [], 'redirect_chain': []
            }
            for attempt in range(1 + retry_count):
                r = self._test_page_with_browser(b, url)
                if r['status'] in ('SUCCESS', 'REDIRECT', 'WARNING'):
                    return r
                last = r
                if attempt < retry_count:
                    time.sleep(retry_delay)
            return last

        # ① Direct
        results['direct']['playwright'] = _test_with_retry(browser)

        # ② Proxy
        proxy_enabled = self.config['proxy']['enabled']
        if proxy_enabled and proxy_browser is not None:
            results['proxy']['playwright'] = _test_with_retry(proxy_browser)
        else:
            results['proxy']['playwright'] = {
                'method': None, 'status': 'SKIPPED', 'http_code': None,
                'response_time': None, 'error_message': 'Proxy disabled',
                'final_url': None, 'redirect_chain': [],
                'console_errors': [], 'network_errors': []
            }

        results['analysis'] = self._analyze_results(results)
        return results

    def test_single_url(self, url: str, url_idx: int = 0, url_total: int = 0) -> Dict:
        """
        하나의 URL에 대해 모든 변형(http/https/www)을 자동 생성하고
        브라우저 1개를 열어 순차 테스트 후 결과를 출력
        (sync_playwright 스레드 충돌 방지를 위해 순차 처리)
        """
        url = self._normalize_url(url)
        variations = self._generate_url_variations(url)
        proxy_enabled = self.config['proxy']['enabled']

        scenarios_per_var = 2 if proxy_enabled else 1
        total_tests = len(variations) * scenarios_per_var

        W = 68
        progress_str = f"  [{url_idx}/{url_total}]" if url_total > 0 else ""
        url_short = url if len(url) <= 50 else url[:47] + '...'

        # ── 헤더 출력 ──
        self._log(f"\n{Fore.CYAN}┌{'─'*(W-2)}┐")
        self._log(f"{Fore.CYAN}│{Style.RESET_ALL}  🔍 {Fore.WHITE}{url_short}{Fore.CYAN}{progress_str}")
        self._log(f"{Fore.CYAN}│{Style.RESET_ALL}    {Fore.WHITE}{len(variations)}개 변형 × {scenarios_per_var}개 시나리오 = {total_tests}개 테스트")
        self._log(f"{Fore.CYAN}├{'─'*(W-2)}┤")
        self._log(f"{Fore.CYAN}│{Style.RESET_ALL}    {'ROUTE':<7}  {'STATUS':<12}  {'TIME':>6}    URL VARIANT")
        self._log(f"{Fore.CYAN}├{'─'*(W-2)}┤")

        variation_results = {}

        # ── 브라우저로 순차 테스트 (sync_playwright 충돌 없음) ──
        browser_args = ['--disable-blink-features=AutomationControlled']
        proxy_config = self._get_playwright_proxy_config() if proxy_enabled else None

        try:
            with sync_playwright() as p:
                # Direct 브라우저
                browser = p.chromium.launch(
                    headless=self.config['headless'],
                    args=browser_args
                )
                # Proxy 브라우저 (enabled일 때만, 같은 playwright 컨텍스트 내에서 안전)
                proxy_browser = None
                if proxy_enabled and proxy_config:
                    try:
                        proxy_browser = p.chromium.launch(
                            headless=self.config['headless'],
                            proxy=proxy_config,
                            args=browser_args
                        )
                    except Exception as pe:
                        self._log(f"  {Fore.YELLOW}▲  Proxy 브라우저 실행 실패: {str(pe)}")
                        proxy_browser = None

                try:
                    for idx, var_url in enumerate(variations, 1):
                        try:
                            var_result = self._test_single_variation(browser, proxy_browser, var_url)
                        except Exception as e:
                            err = {
                                'method': None, 'status': 'ERROR', 'http_code': None,
                                'response_time': None, 'error_message': str(e),
                                'final_url': None, 'redirect_chain': [],
                                'console_errors': [], 'network_errors': []
                            }
                            proxy_err = {
                                **err,
                                'status': 'ERROR' if proxy_enabled else 'SKIPPED',
                                'error_message': str(e) if proxy_enabled else 'Proxy disabled',
                            }
                            var_result = {
                                'url': var_url,
                                'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                                'direct': {'playwright': err},
                                'proxy': {'playwright': proxy_err},
                                'analysis': {'direct_vs_proxy': None, 'final_verdict': 'ERROR', 'issues': [str(e)]}
                            }

                        variation_results[var_url] = var_result
                        dp = var_result['direct']['playwright']
                        pp = var_result['proxy']['playwright']
                        verdict = var_result['analysis']['final_verdict']

                        url_display = var_url if len(var_url) <= 55 else var_url[:52] + '...'
                        self._log(f"  {Fore.WHITE}│  {Fore.CYAN}[{idx}/{len(variations)}]{Style.RESET_ALL} {url_display}")
                        self._log(
                            f"  {Fore.WHITE}│    {Fore.BLUE}DIRECT {Style.RESET_ALL}"
                            f"  {self._get_status_badge(dp['status'])}  {Fore.WHITE}{self._fmt_time(dp['response_time'])}"
                        )
                        if proxy_enabled:
                            self._log(
                                f"  {Fore.WHITE}│    {Fore.MAGENTA}PROXY  {Style.RESET_ALL}"
                                f"  {self._get_status_badge(pp['status'])}  {Fore.WHITE}{self._fmt_time(pp['response_time'])}"
                            )
                        else:
                            self._log(f"  {Fore.WHITE}│    {Fore.YELLOW}PROXY  {Style.RESET_ALL}  {self._get_status_badge('SKIPPED')}       -")
                        self._log(f"  {Fore.WHITE}│    {'':7}  {self._get_status_badge(verdict)}  {Fore.WHITE}판정")

                        if idx < len(variations):
                            self._log(f"  {Fore.WHITE}│{'·'*(W-4)}")
                finally:
                    browser.close()
                    if proxy_browser:
                        proxy_browser.close()

        except Exception as e:
            # 브라우저 실행 자체 실패
            self._log(f"  {Fore.RED}✕  브라우저 실행 실패: {str(e)}")
            err = {
                'method': None, 'status': 'ERROR', 'http_code': None,
                'response_time': None, 'error_message': f'Browser launch error: {str(e)}',
                'final_url': None, 'redirect_chain': [], 'console_errors': [], 'network_errors': []
            }
            for var_url in variations:
                variation_results[var_url] = {
                    'url': var_url,
                    'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                    'direct': {'playwright': err},
                    'proxy': {'playwright': {**err, 'status': 'SKIPPED', 'error_message': 'Proxy disabled'}},
                    'analysis': {'direct_vs_proxy': None, 'final_verdict': 'ERROR', 'issues': [str(e)]}
                }

        # ── 종합 판정 ──
        overall = self._analyze_all_variations(variation_results)

        self._log(f"{Fore.CYAN}├{'─'*(W-2)}┤")
        verdict_badge = self._get_status_badge(overall['final_verdict'])
        self._log(f"{Fore.CYAN}│{Style.RESET_ALL}  종합 판정  {verdict_badge}")

        if overall['issues']:
            for issue in overall['issues'][:3]:
                issue_short = issue if len(issue) <= W-8 else issue[:W-11] + '...'
                self._log(f"{Fore.CYAN}│{Style.RESET_ALL}  {Fore.YELLOW}⚠  {issue_short}")
        self._log(f"{Fore.CYAN}└{'─'*(W-2)}┘")

        return {
            'url': url,
            'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'variations': variation_results,
            'analysis': overall
        }

    def _analyze_all_variations(self, variation_results: Dict[str, Dict]) -> Dict:
        """모든 URL 변형의 결과를 종합 분석"""
        analysis = {
            'final_verdict': None,
            'issues': [],
            'variation_verdicts': {}
        }

        verdicts = []
        for var_url, var_result in variation_results.items():
            v = var_result.get('analysis', {}).get('final_verdict', 'ERROR')
            analysis['variation_verdicts'][var_url] = v
            verdicts.append(v)

        # SKIPPED 제외한 유효 상태만으로 판정
        effective_verdicts = [v for v in verdicts if v != 'SKIPPED']

        # 하나라도 성공이면 전체 SUCCESS
        if any(v == 'SUCCESS' for v in effective_verdicts):
            analysis['final_verdict'] = 'SUCCESS'
        elif any(v == 'REDIRECT' for v in effective_verdicts):
            analysis['final_verdict'] = 'REDIRECT'
        elif any(v == 'WARNING' for v in effective_verdicts):
            analysis['final_verdict'] = 'WARNING'
        elif effective_verdicts and all(v == 'TIMEOUT' for v in effective_verdicts):
            analysis['final_verdict'] = 'TIMEOUT'
        elif not effective_verdicts:
            analysis['final_verdict'] = 'SKIPPED'
        else:
            analysis['final_verdict'] = 'ERROR'

        # 변형별 차이 분석
        success_vars = [u for u, v in analysis['variation_verdicts'].items() if v == 'SUCCESS']
        error_vars = [u for u, v in analysis['variation_verdicts'].items() if v == 'ERROR']

        if success_vars and error_vars:
            analysis['issues'].append(
                f"일부 변형만 성공: {', '.join(success_vars)}"
            )

        # http vs https 차이
        https_results = {u: v for u, v in analysis['variation_verdicts'].items() if u.startswith('https://')}
        http_results = {u: v for u, v in analysis['variation_verdicts'].items() if u.startswith('http://') and not u.startswith('https://')}

        https_ok = any(v in ('SUCCESS', 'REDIRECT') for v in https_results.values())
        http_ok = any(v in ('SUCCESS', 'REDIRECT') for v in http_results.values())

        if https_ok and not http_ok:
            analysis['issues'].append('HTTP 접속 불가 - HTTPS만 지원')
        elif http_ok and not https_ok:
            analysis['issues'].append('HTTPS 접속 불가 - HTTP만 지원')

        # www 차이
        www_results = {u: v for u, v in analysis['variation_verdicts'].items() if '://www.' in u}
        non_www_results = {u: v for u, v in analysis['variation_verdicts'].items() if '://www.' not in u}

        www_ok = any(v in ('SUCCESS', 'REDIRECT') for v in www_results.values())
        non_www_ok = any(v in ('SUCCESS', 'REDIRECT') for v in non_www_results.values())

        if www_ok and not non_www_ok:
            analysis['issues'].append('www 없이 접속 불가 - www 필수')
        elif non_www_ok and not www_ok:
            analysis['issues'].append('www 접속 불가 - www 없이만 가능')

        # 개별 변형 이슈 수집
        for var_url, var_result in variation_results.items():
            for issue in var_result.get('analysis', {}).get('issues', []):
                prefixed = f"[{var_url}] {issue}"
                if prefixed not in analysis['issues']:
                    analysis['issues'].append(prefixed)

        return analysis

    # 상태별 표시 설정 (색상, 아이콘, 레이블) — 폭 고정(10자)으로 컬럼 정렬
    _STATUS_STYLE = {
        'SUCCESS':  (Fore.GREEN,   '●', 'SUCCESS '),
        'REDIRECT': (Fore.CYAN,    '↪', 'REDIRECT'),
        'WARNING':  (Fore.YELLOW,  '▲', 'WARNING '),
        'TIMEOUT':  (Fore.YELLOW,  '◷', 'TIMEOUT '),
        'ERROR':    (Fore.RED,     '✕', 'ERROR   '),
        'SKIPPED':  (Fore.WHITE,   '─', 'SKIPPED '),
    }

    def _get_status_badge(self, status: str) -> str:
        """고정폭 상태 배지 반환 (컬럼 정렬용)"""
        color, icon, label = self._STATUS_STYLE.get(
            status, (Fore.WHITE, '?', 'UNKNOWN ')
        )
        return f"{color}[{icon} {label}]{Style.RESET_ALL}"

    def _analyze_results(self, results: Dict) -> Dict:
        """2가지 테스트 결과 종합 분석"""
        analysis = {
            'direct_vs_proxy': None,
            'final_verdict': None,
            'issues': []
        }

        # Direct 결과 (None 방어)
        dp_pw = (results.get('direct') or {}).get('playwright') or {}
        dp_status = dp_pw.get('status') or 'ERROR'

        # Proxy 결과 (None 방어)
        pp_pw = (results.get('proxy') or {}).get('playwright') or {}
        pp_status = pp_pw.get('status') or 'SKIPPED'

        # 판정에 사용할 상태 목록 (SKIPPED 제외)
        all_statuses = []
        if dp_status != 'SKIPPED':
            all_statuses.append(dp_status)
        if pp_status != 'SKIPPED':
            all_statuses.append(pp_status)

        # 최종 판정
        if not all_statuses:
            analysis['final_verdict'] = 'SKIPPED'
        elif any(s == 'SUCCESS' for s in all_statuses):
            analysis['final_verdict'] = 'SUCCESS'
        elif any(s == 'REDIRECT' for s in all_statuses):
            analysis['final_verdict'] = 'REDIRECT'
        elif any(s == 'WARNING' for s in all_statuses):
            analysis['final_verdict'] = 'WARNING'
        elif all(s == 'TIMEOUT' for s in all_statuses):
            analysis['final_verdict'] = 'TIMEOUT'
        else:
            analysis['final_verdict'] = 'ERROR'

        # 문제점 분석
        if pp_status != 'SKIPPED':
            if dp_status == 'SUCCESS' and pp_status == 'ERROR':
                analysis['issues'].append('Proxy connection issue')

        if dp_pw.get('console_errors'):
            count = len(dp_pw['console_errors'])
            analysis['issues'].append(f'Direct: {count} console errors')

        if pp_status != 'SKIPPED' and pp_pw.get('console_errors'):
            count = len(pp_pw['console_errors'])
            analysis['issues'].append(f'Proxy: {count} console errors')

        return analysis

    def test_multiple_urls(self, urls: List[str]) -> List[Dict]:
        """
        여러 URL을 순차 테스트 (URL당 sync_playwright 1회 / 브라우저 1~2개 재사용)

        Args:
            urls: URL 리스트

        Returns:
            전체 테스트 결과 리스트 (sites.txt 순서 유지)
        """
        total = len(urls)
        proxy_enabled = self.config['proxy']['enabled']
        scenarios = 2 if proxy_enabled else 1
        max_tests = total * 4 * scenarios  # 최대값 (URL 중복 변형 제거 시 실제는 더 적을 수 있음)

        W = 68
        self._log(f"\n{Fore.CYAN}┌{'─'*(W-2)}┐")
        self._log(f"{Fore.CYAN}│{Style.RESET_ALL}  실행 계획")
        self._log(f"{Fore.CYAN}├{'─'*(W-2)}┤")
        self._log(f"{Fore.CYAN}│{Style.RESET_ALL}  모드       : {Fore.WHITE}URL 순차 처리 / 변형 순차 (브라우저 1개 재사용)")
        self._log(f"{Fore.CYAN}│{Style.RESET_ALL}  대상 URL   : {Fore.WHITE}{total}개")
        self._log(f"{Fore.CYAN}│{Style.RESET_ALL}  최대 테스트: {Fore.WHITE}{total} URL × 최대 4 변형 × {scenarios} 시나리오 = {max_tests}건")
        self._log(f"{Fore.CYAN}└{'─'*(W-2)}┘\n")

        # URL은 순차 처리 (출력 섞임 방지) — 변형 병렬화는 test_single_url 내부에서 수행
        results = []
        for idx, url in enumerate(urls, 1):
            try:
                result = self.test_single_url(url, url_idx=idx, url_total=total)
                result['order'] = idx
            except Exception as e:
                self._log(f"  {Fore.RED}✕  {url} 테스트 실패: {e}")
                result = {
                    'url': url,
                    'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                    'variations': {},
                    'analysis': {'final_verdict': 'ERROR', 'issues': [str(e)], 'variation_verdicts': {}},
                    'order': idx,
                }
            results.append(result)

        self.results = results
        return results

    def _safe_open_csv(self, file_path: str):
        """CSV 파일을 안전하게 열기 (PermissionError 대응)"""
        try:
            f = open(file_path, 'w', newline='', encoding='utf-8-sig')
            return f
        except PermissionError:
            self._log(f"  {Fore.RED}✕  '{file_path}' 파일에 접근할 수 없습니다.")
            self._log(f"  {Fore.YELLOW}   → Excel이나 다른 프로그램에서 파일을 닫아주세요.")

            base, ext = os.path.splitext(file_path)
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            alt_path = f"{base}_{timestamp}{ext}"

            self._log(f"  {Fore.CYAN}   → 대체 파일로 저장 시도: {alt_path}")
            try:
                f = open(alt_path, 'w', newline='', encoding='utf-8-sig')
                return f
            except PermissionError:
                self._log(f"  {Fore.RED}✕  대체 파일도 실패. 저장을 건너뜁니다.")
                return None

    @staticmethod
    def _pick_best_result(result: Dict):
        """
        한 URL의 모든 변형/시나리오 중 가장 좋은 결과 하나를 선택.
        우선순위: SUCCESS > REDIRECT > WARNING > TIMEOUT > ERROR
        성공 결과가 없으면 None 반환 (→ 원본 URL + 전체 판정 상태 사용)
        """
        priority = {'SUCCESS': 0, 'REDIRECT': 1, 'WARNING': 2, 'TIMEOUT': 3, 'ERROR': 4, 'SKIPPED': 5}
        best = None
        best_score = 99

        for var_url, var_result in result['variations'].items():
            for key in ('direct', 'proxy'):
                pw = (var_result.get(key) or {}).get('playwright') or {}
                status = pw.get('status') or 'ERROR'
                score = priority.get(status, 99)
                if score < best_score:
                    best_score = score
                    best = {
                        'url': var_url,
                        'status': status,
                        'http_code': pw.get('http_code') or '',
                        'response_time': pw.get('response_time'),
                        'final_url': pw.get('final_url') or var_url,
                        'source': key,
                    }
        return best

    def save_to_xlsx(self, output_file: str = 'results.xlsx'):
        """결과를 XLSX로 저장 — URL당 1행, 접속 가능 URL 우선 표시"""
        if not self.results:
            self._log(f"  {Fore.YELLOW}▲  저장할 결과가 없습니다.")
            return

        # openpyxl 없으면 CSV fallback
        if not _OPENPYXL_AVAILABLE:
            self._log(f"  {Fore.YELLOW}▲  openpyxl 미설치 → CSV로 저장합니다. (pip install openpyxl)")
            self._save_summary_csv(output_file.replace('.xlsx', '.csv'))
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '결과'

        # ── 색상 정의 ──
        fill = {
            'SUCCESS':  PatternFill('solid', fgColor='C6EFCE'),
            'REDIRECT': PatternFill('solid', fgColor='DDEBF7'),
            'WARNING':  PatternFill('solid', fgColor='FFEB9C'),
            'TIMEOUT':  PatternFill('solid', fgColor='FFEB9C'),
            'ERROR':    PatternFill('solid', fgColor='FFC7CE'),
        }
        bold = Font(bold=True)
        center = Alignment(horizontal='center', vertical='center')

        # ── 헤더 ──
        headers = ['No', 'Original_URL', 'Accessible_URL', '접속가능',
                   'Status', 'Response_Time(s)', 'Source', 'Timestamp', 'Issues']
        ws.append(headers)
        for cell in ws[1]:
            cell.font = bold
            cell.alignment = center

        # ── 데이터 행 (URL당 1행) ──
        for result in self.results:
            best = self._pick_best_result(result)
            overall_verdict = result['analysis']['final_verdict']
            issues = '; '.join(result['analysis']['issues']) if result['analysis']['issues'] else ''
            timestamp = result.get('timestamp', '')

            if best and best['status'] in ('SUCCESS', 'REDIRECT', 'WARNING'):
                accessible_url = best['final_url'] or best['url']
                status = best['status']
                reachable = 'O'
                resp_time = f"{best['response_time']:.3f}" if best['response_time'] else ''
                source = best['source']
            else:
                # 접속 불가 → 원본 URL 그대로
                accessible_url = result['url']
                status = overall_verdict
                reachable = 'X'
                resp_time = ''
                source = ''

            row = [
                result.get('order', ''),
                result['url'],
                accessible_url,
                reachable,
                status,
                resp_time,
                source,
                timestamp,
                issues,
            ]
            ws.append(row)

            # 상태에 맞는 행 색상 적용
            row_fill = fill.get(status, fill['ERROR'])
            for cell in ws[ws.max_row]:
                cell.fill = row_fill

        # ── 열 너비 자동 조정 ──
        col_widths = [6, 45, 55, 10, 12, 16, 8, 20, 60]
        for i, width in enumerate(col_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

        # ── 저장 ──
        try:
            wb.save(output_file)
            self._log(f"  {Fore.GREEN}●  XLSX 저장 완료 : {output_file}  ({len(self.results)}행)")
        except PermissionError:
            base, ext = os.path.splitext(output_file)
            alt = f"{base}_{datetime.now().strftime('%Y%m%d_%H%M%S')}{ext}"
            self._log(f"  {Fore.YELLOW}▲  파일 접근 불가 → 대체 저장: {alt}")
            try:
                wb.save(alt)
                self._log(f"  {Fore.GREEN}●  XLSX 저장 완료 : {alt}")
            except Exception as e2:
                self._log(f"  {Fore.RED}✕  대체 파일 저장도 실패: {e2}")

    def _save_summary_csv(self, output_file: str):
        """openpyxl 없을 때 CSV fallback (URL당 1행)"""
        headers = ['No', 'Original_URL', 'Accessible_URL', '접속가능',
                   'Status', 'Response_Time(s)', 'Source', 'Timestamp', 'Issues']
        f = self._safe_open_csv(output_file)
        if not f:
            return
        try:
            writer = csv.DictWriter(f, fieldnames=headers)
            writer.writeheader()
            for result in self.results:
                best = self._pick_best_result(result)
                overall_verdict = result['analysis']['final_verdict']
                issues = '; '.join(result['analysis']['issues']) if result['analysis']['issues'] else ''
                if best and best['status'] in ('SUCCESS', 'REDIRECT', 'WARNING'):
                    accessible_url = best['final_url'] or best['url']
                    status = best['status']
                    reachable = 'O'
                    resp_time = f"{best['response_time']:.3f}" if best['response_time'] else ''
                    source = best['source']
                else:
                    accessible_url = result['url']
                    status = overall_verdict
                    reachable = 'X'
                    resp_time = ''
                    source = ''
                writer.writerow({
                    'No': result.get('order', ''),
                    'Original_URL': result['url'],
                    'Accessible_URL': accessible_url,
                    '접속가능': reachable,
                    'Status': status,
                    'Response_Time(s)': resp_time,
                    'Source': source,
                    'Timestamp': result.get('timestamp', ''),
                    'Issues': issues,
                })
            self._log(f"  {Fore.GREEN}●  CSV 저장 완료  : {f.name}")
        finally:
            f.close()

    def print_summary(self):
        """테스트 결과 요약 출력"""
        if not self.results:
            self._log(f"{Fore.YELLOW}  ⚠  결과가 없습니다.")
            return

        W = 68
        total = len(self.results)
        counts = {
            'SUCCESS':  sum(1 for r in self.results if r['analysis']['final_verdict'] == 'SUCCESS'),
            'REDIRECT': sum(1 for r in self.results if r['analysis']['final_verdict'] == 'REDIRECT'),
            'WARNING':  sum(1 for r in self.results if r['analysis']['final_verdict'] == 'WARNING'),
            'ERROR':    sum(1 for r in self.results if r['analysis']['final_verdict'] == 'ERROR'),
            'TIMEOUT':  sum(1 for r in self.results if r['analysis']['final_verdict'] == 'TIMEOUT'),
        }
        def pct(n: int) -> str:
            return f"{n/total*100:.1f}%" if total > 0 else "0.0%"
        BAR_W = 20

        def bar(n: int) -> str:
            filled = round(BAR_W * n / total) if total > 0 else 0
            return '█' * filled + '░' * (BAR_W - filled)

        self._log(f"\n{Fore.CYAN}┌{'─'*(W-2)}┐")
        self._log(f"{Fore.CYAN}│{Style.RESET_ALL}  테스트 결과 요약")
        self._log(f"{Fore.CYAN}├{'─'*(W-2)}┤")
        self._log(f"{Fore.CYAN}│{Style.RESET_ALL}  총 사이트 수 : {Fore.WHITE}{total}개")
        self._log(f"{Fore.CYAN}├{'─'*(W-2)}┤")
        self._log(f"{Fore.CYAN}│{Style.RESET_ALL}  {'상태':<8}  {'건수':>4}  {'비율':>6}  분포")
        self._log(f"{Fore.CYAN}├{'─'*(W-2)}┤")

        for status, (color, icon, label) in self._STATUS_STYLE.items():
            if status == 'SKIPPED':
                continue
            n = counts.get(status, 0)
            b = bar(n)
            label_str = f"{color}{icon} {label.strip():<9}{Style.RESET_ALL}"
            self._log(
                f"{Fore.CYAN}│{Style.RESET_ALL}  {label_str} {n:>4}  {pct(n):>6}  "
                f"{color}{b}{Style.RESET_ALL}"
            )

        self._log(f"{Fore.CYAN}└{'─'*(W-2)}┘")

        # ── 사이트별 상세 ──
        self._log(f"\n{Fore.CYAN}┌{'─'*(W-2)}┐")
        self._log(f"{Fore.CYAN}│{Style.RESET_ALL}  사이트별 상세")
        self._log(f"{Fore.CYAN}├{'─'*(W-2)}┤")

        for result in self.results:
            verdict = result['analysis']['final_verdict']
            order = result.get('order', '-')
            url_short = result['url'] if len(result['url']) <= 50 else result['url'][:47] + '...'
            badge = self._get_status_badge(verdict)
            self._log(f"{Fore.CYAN}│{Style.RESET_ALL}  {Fore.WHITE}{order:>3}.{Style.RESET_ALL}  {badge}  {Fore.WHITE}{url_short}")

            for var_url, v in result['analysis'].get('variation_verdicts', {}).items():
                var_short = var_url if len(var_url) <= 52 else var_url[:49] + '...'
                self._log(f"{Fore.CYAN}│{Style.RESET_ALL}         {self._get_status_badge(v)}  {Fore.WHITE}{Style.DIM}{var_short}{Style.RESET_ALL}")

            if result['analysis']['issues']:
                for issue in result['analysis']['issues'][:3]:
                    issue_short = issue if len(issue) <= W-12 else issue[:W-15] + '...'
                    self._log(f"{Fore.CYAN}│{Style.RESET_ALL}         {Fore.YELLOW}⚠  {issue_short}{Style.RESET_ALL}")

            self._log(f"{Fore.CYAN}│{Style.RESET_ALL}")

        self._log(f"{Fore.CYAN}└{'─'*(W-2)}┘\n")


def load_urls_from_file(file_path: str) -> List[str]:
    """텍스트 파일에서 URL 목록 로드"""
    urls = []
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith('#'):
                    normalized = URLAliveCheck._normalize_url(line)
                    if normalized:
                        urls.append(normalized)
    except FileNotFoundError:
        print(f"  {Fore.RED}✕  파일을 찾을 수 없습니다: {file_path}")
    except Exception as e:
        print(f"  {Fore.RED}✕  파일 읽기 오류: {e}")

    return urls


# ── 공유폴더 UNC 경로 설정 ──────────────────────────────────────
SHARED_FOLDER = r'\\server\share\monitor'
# ────────────────────────────────────────────────────────────────


def find_local_file(filename: str) -> Optional[str]:
    """로컬(스크립트 폴더 / 작업 디렉토리)에서 파일 탐색"""
    script_dir = os.path.dirname(os.path.abspath(__file__))
    for base in (script_dir, os.getcwd()):
        path = os.path.join(base, filename)
        if os.path.isfile(path):
            return path
    return None


def find_default_file(filename: str) -> Optional[str]:
    """
    파일 탐색 우선순위:
      1. 공유폴더(SHARED_FOLDER)
      2. 공유폴더 실패 시 → 로컬 사용 여부 사용자에게 질문
    출력은 박스 내부 라인 형식(│로 시작)으로 통일
    """
    # ── 1) 공유폴더 시도 ──
    shared_path = os.path.join(SHARED_FOLDER, filename)
    try:
        if os.path.isfile(shared_path):
            print(f"{Fore.CYAN}│{Style.RESET_ALL}    {Fore.GREEN}●{Style.RESET_ALL}  [{filename}] 공유폴더에서 로드: {shared_path}")
            return shared_path
        else:
            print(f"{Fore.CYAN}│{Style.RESET_ALL}    {Fore.YELLOW}▲{Style.RESET_ALL}  [{filename}] 공유폴더에 없음")
    except Exception as e:
        print(f"{Fore.CYAN}│{Style.RESET_ALL}    {Fore.YELLOW}▲{Style.RESET_ALL}  [{filename}] 공유폴더 접근 실패: {e}")

    # ── 2) 로컬 fallback ──
    local_path = find_local_file(filename)
    if local_path:
        print(f"{Fore.CYAN}│{Style.RESET_ALL}    {Fore.YELLOW}▲{Style.RESET_ALL}  [{filename}] 로컬에서 발견: {local_path}")
        answer = input(f"{Fore.CYAN}│{Style.RESET_ALL}       로컬 파일을 사용할까요? [Y/n]: ").strip().lower()
        if answer in ('', 'y', 'yes'):
            return local_path
        else:
            print(f"{Fore.CYAN}│{Style.RESET_ALL}    {Fore.RED}✕{Style.RESET_ALL}  [{filename}] 로컬 파일 사용 취소.")
            return None
    else:
        print(f"{Fore.CYAN}│{Style.RESET_ALL}    {Fore.RED}✕{Style.RESET_ALL}  [{filename}] 로컬에도 없음.")
        return None


def main():
    """메인 실행 함수"""
    parser = argparse.ArgumentParser(
        description='URL_Alive_Check - URL 접속 상태 확인 도구',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
예시:
  # 인자 없이 실행 (자동으로 sites.txt + config.json 감지)
  python url_alive_check.py

  # 파일 지정
  python url_alive_check.py --input sites.txt --output results.xlsx

  # 단일 URL 테스트
  python url_alive_check.py --url naver.com
        """
    )

    # 실행 위치 기준 기본 출력 경로 (날짜 포함)
    timestamp = datetime.now().strftime('%y%m%d')
    default_output = os.path.join(os.getcwd(), f'results_{timestamp}.xlsx')

    parser.add_argument('--input', '-i', help='URL 목록 파일 경로 (기본: sites.txt 자동 감지)')
    parser.add_argument('--output', '-o', default=default_output, help='결과 XLSX 파일 경로')
    parser.add_argument('--config', '-c', help='설정 파일 경로 (기본: config.json 자동 감지)')
    parser.add_argument('--url', '-u', help='단일 URL 테스트')
    parser.add_argument('--sequential', '-s', action='store_true', help='순차 실행 모드 (기본 동작과 동일, 명시적 지정용)')

    args = parser.parse_args()

    W = 68
    now_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    # ── 배너 (가장 먼저 출력) ──
    print(f"\n{Fore.CYAN}╔{'═'*(W-2)}╗")
    print(f"{Fore.CYAN}║")
    print(f"{Fore.CYAN}║{Style.RESET_ALL}  {Fore.WHITE}URL_Alive_Check  v2.1")
    print(f"{Fore.CYAN}║{Style.RESET_ALL}  {Fore.WHITE}URL 접속 상태 확인")
    print(f"{Fore.CYAN}║")
    print(f"{Fore.CYAN}║{Style.RESET_ALL}  {Style.DIM}URL → http/https × www/non-www × Direct/Proxy{Style.RESET_ALL}")
    print(f"{Fore.CYAN}║{Style.RESET_ALL}  {Style.DIM}{now_str}{Style.RESET_ALL}")
    print(f"{Fore.CYAN}║")
    print(f"{Fore.CYAN}╚{'═'*(W-2)}╝\n")

    # ── 옵션 안내 박스 ──
    print(f"{Fore.YELLOW}┌{'─'*(W-2)}┐")
    print(f"{Fore.YELLOW}│{Style.RESET_ALL}  사용 가능한 옵션")
    print(f"{Fore.YELLOW}├{'─'*(W-2)}┤")
    print(f"{Fore.YELLOW}│{Style.RESET_ALL}  {Fore.WHITE}-i / --input{Style.RESET_ALL}   <파일>   URL 목록 파일 지정  {Style.DIM}(기본: sites.txt 자동 감지){Style.RESET_ALL}")
    print(f"{Fore.YELLOW}│{Style.RESET_ALL}  {Fore.WHITE}-o / --output{Style.RESET_ALL}  <파일>   결과 XLSX 저장 경로 {Style.DIM}(기본: results_YYMMDD.xlsx){Style.RESET_ALL}")
    print(f"{Fore.YELLOW}│{Style.RESET_ALL}  {Fore.WHITE}-c / --config{Style.RESET_ALL}  <파일>   설정 파일 지정     {Style.DIM}(기본: config.json 자동 감지){Style.RESET_ALL}")
    print(f"{Fore.YELLOW}│{Style.RESET_ALL}  {Fore.WHITE}-u / --url{Style.RESET_ALL}     <URL>    단일 URL 즉시 테스트")
    print(f"{Fore.YELLOW}├{'─'*(W-2)}┤")
    print(f"{Fore.YELLOW}│{Style.RESET_ALL}  {Style.DIM}예시: python url_alive_check.py -i sites.txt -o results.xlsx{Style.RESET_ALL}")
    print(f"{Fore.YELLOW}│{Style.RESET_ALL}  {Style.DIM}예시: python url_alive_check.py -u https://example.com{Style.RESET_ALL}")
    print(f"{Fore.YELLOW}└{'─'*(W-2)}┘\n")

    # ── 환경 정보 박스 시작 ──
    try:
        shared_ok = os.path.isdir(SHARED_FOLDER)
    except Exception:
        shared_ok = False

    print(f"{Fore.CYAN}┌{'─'*(W-2)}┐")
    print(f"{Fore.CYAN}│{Style.RESET_ALL}  환경 설정")
    print(f"{Fore.CYAN}├{'─'*(W-2)}┤")

    shared_status = f"{Fore.GREEN}접근 가능{Style.RESET_ALL}" if shared_ok else f"{Fore.YELLOW}접근 불가 → 로컬 fallback{Style.RESET_ALL}"
    print(f"{Fore.CYAN}│{Style.RESET_ALL}  공유폴더   : {shared_status}")

    # config.json 탐색 (박스 내부에서 출력)
    config_file = args.config
    if not config_file:
        config_file = find_default_file('config.json')
    config_status = f"{Fore.GREEN}{config_file}{Style.RESET_ALL}" if config_file else f"{Fore.YELLOW}없음 (기본값 사용){Style.RESET_ALL}"
    print(f"{Fore.CYAN}│{Style.RESET_ALL}  설정 파일  : {config_status}")

    # sites.txt 탐색 (박스 내부에서 출력)
    input_file = args.input
    if not input_file and not args.url:
        input_file = find_default_file('sites.txt')

    sites_status = f"{Fore.GREEN}{input_file}{Style.RESET_ALL}" if input_file else (
        f"{Fore.WHITE}{args.url}{Style.RESET_ALL}  (단일 URL)" if args.url else f"{Fore.RED}없음{Style.RESET_ALL}"
    )
    print(f"{Fore.CYAN}│{Style.RESET_ALL}  URL 목록   : {sites_status}")

    print(f"{Fore.CYAN}└{'─'*(W-2)}┘\n")

    # sites.txt 없으면 여기서 즉시 종료 (Enter 대기 전에)
    if not args.url and not input_file:
        print(f"  {Fore.RED}✕  sites.txt를 찾을 수 없습니다.")
        print(f"  {Fore.YELLOW}     다음 중 하나를 실행하세요:")
        print(f"         1.  같은 폴더에 sites.txt 파일 생성")
        print(f"         2.  python url_alive_check.py --input <파일경로>")
        print(f"         3.  python url_alive_check.py --url <URL>\n")
        return

    # 모니터 객체 생성
    monitor = URLAliveCheck(config_file=config_file)

    # Playwright 사전 검증
    if not URLAliveCheck.check_playwright_available():
        print(f"  {Fore.RED}✕  Playwright 검증 실패. 프로그램을 종료합니다.")
        sys.exit(1)

    # URL 로드
    urls = []
    if args.url:
        normalized = URLAliveCheck._normalize_url(args.url)
        if not normalized:
            print(f"  {Fore.RED}✕  유효하지 않은 URL입니다: '{args.url}'\n")
            return
        urls = [normalized]
    else:
        urls = load_urls_from_file(input_file)

    if not urls:
        print(f"\n{Fore.RED}  ✕  테스트할 URL이 없습니다.\n")
        return

    # 실행 정보 요약 출력
    url_count = len(urls)
    proxy_cfg = monitor.config.get('proxy', {})
    proxy_enabled = proxy_cfg.get('enabled', False)

    print(f"{Fore.CYAN}┌{'─'*(W-2)}┐")
    print(f"{Fore.CYAN}│{Style.RESET_ALL}  실행 정보")
    print(f"{Fore.CYAN}├{'─'*(W-2)}┤")
    if args.url:
        print(f"{Fore.CYAN}│{Style.RESET_ALL}  입력 URL   : {Fore.WHITE}{args.url}")
    else:
        print(f"{Fore.CYAN}│{Style.RESET_ALL}  URL 파일   : {Fore.WHITE}{input_file}  ({url_count}개)")
    print(f"{Fore.CYAN}│{Style.RESET_ALL}  출력 파일  : {Fore.WHITE}{args.output}")

    # ── URL 목록 미리보기 ──
    print(f"{Fore.CYAN}├{'─'*(W-2)}┤")
    print(f"{Fore.CYAN}│{Style.RESET_ALL}  URL 목록 ({url_count}개)")
    PREVIEW_MAX = 10
    for i, u in enumerate(urls[:PREVIEW_MAX], 1):
        u_short = u if len(u) <= W - 10 else u[:W - 13] + '...'
        print(f"{Fore.CYAN}│{Style.RESET_ALL}    {Fore.WHITE}{i:>3}.{Style.RESET_ALL}  {u_short}")
    if url_count > PREVIEW_MAX:
        print(f"{Fore.CYAN}│{Style.RESET_ALL}         {Style.DIM}... 외 {url_count - PREVIEW_MAX}개{Style.RESET_ALL}")

    # ── 프록시 설정 ──
    print(f"{Fore.CYAN}├{'─'*(W-2)}┤")
    if proxy_enabled:
        proxy_server = proxy_cfg.get('server', '-')
        use_auth = proxy_cfg.get('use_auth', False)
        proxy_user = proxy_cfg.get('username', '')
        auth_str = f"  인증: {proxy_user}" if use_auth and proxy_user else ("  인증: 없음" if not use_auth else "  인증: 설정됨")
        print(f"{Fore.CYAN}│{Style.RESET_ALL}  프록시     : {Fore.GREEN}사용{Style.RESET_ALL}  {proxy_server}")
        print(f"{Fore.CYAN}│{Style.RESET_ALL}             {Style.DIM}{auth_str}{Style.RESET_ALL}")
    else:
        print(f"{Fore.CYAN}│{Style.RESET_ALL}  프록시     : {Fore.YELLOW}미사용{Style.RESET_ALL}")

    print(f"{Fore.CYAN}└{'─'*(W-2)}┘\n")

    # 실행 전 확인 대기 (모든 정보 확인 후 Enter)
    print(f"{Fore.CYAN}┌{'─'*(W-2)}┐")
    print(f"{Fore.CYAN}│{Style.RESET_ALL}  {Fore.GREEN}●  준비 완료.{Style.RESET_ALL}  모니터링을 시작하려면 {Fore.WHITE}Enter{Style.RESET_ALL}를 누르세요.")
    print(f"{Fore.CYAN}└{'─'*(W-2)}┘")
    try:
        input()
    except (EOFError, KeyboardInterrupt):
        print(f"\n  {Fore.YELLOW}▲  사용자가 취소했습니다.")
        sys.exit(0)

    start_time = time.time()

    # 테스트 실행
    monitor.test_multiple_urls(urls)

    # 결과 저장
    monitor.save_to_xlsx(args.output)

    # 요약 출력
    monitor.print_summary()

    elapsed_time = time.time() - start_time
    m, s = divmod(int(elapsed_time), 60)
    elapsed_str = f"{m}분 {s:02d}초" if m > 0 else f"{elapsed_time:.2f}초"

    print(f"{Fore.CYAN}┌{'─'*(W-2)}┐")
    print(f"{Fore.CYAN}│{Style.RESET_ALL}  완료")
    print(f"{Fore.CYAN}├{'─'*(W-2)}┤")
    print(f"{Fore.CYAN}│{Style.RESET_ALL}  소요 시간  : {Fore.WHITE}{elapsed_str}")
    print(f"{Fore.CYAN}│{Style.RESET_ALL}  저장 파일  : {Fore.WHITE}{args.output}")
    print(f"{Fore.CYAN}│{Style.RESET_ALL}  {Fore.GREEN}●  모든 작업이 완료되었습니다.{Style.RESET_ALL}")
    print(f"{Fore.CYAN}└{'─'*(W-2)}┘\n")


if __name__ == '__main__':
    main()
